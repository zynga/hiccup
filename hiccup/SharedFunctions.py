# Hiccup - Burp Suite Python Extensions
# Copyright 2012 Zynga Inc.

import re
import json
import array

from openpyxl import Workbook
from openpyxl.style import Alignment
from openpyxl.cell import get_column_letter

def update_json_values(jsonobj, updates, changed):
    if (jsonobj != None):
        if (isinstance(jsonobj,list)):
            for i in jsonobj:
                changed = update_json_values(i, updates, changed)
        elif (isinstance(jsonobj,dict)):
            for (k,v) in jsonobj.items():
                if (k in updates.keys()):
                    jsonobj[k] == updates[k]
                    changed = True
                changed = update_json_values(v, updates, changed)
    return changed

def indent(data, tabs):
    newstr = ''
    for line in data.splitlines():
        newstr = newstr + '\t'*tabs + line + '\n'
    return newstr.rstrip()

def pprint_table(self, table):
    """Prints out a table of data, padded for alignment
    @param table: The table to print. A list of lists.
    Each row must have the same number of columns. """
    col_paddings = []

    for i in range(len(table[0])):
        col_paddings.append(max([len(str(row[i])) for row in table]))

    for row in table:
        # left col
        print row[0].ljust(col_paddings[0] + 1),
        # rest of the cols
        for i in range(1, len(row)):
            col = str(row[i]).rjust(col_paddings[i] + 2)
            print col,
        print

def write_xlsx(filename, worksheet, columns, table):
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = worksheet
    for (col_letter, col_title, col_width) in columns:
        ws.cell("%s1" % (col_letter)).value = col_title
        ws.cell("%s1" % (col_letter)).style.font.bold = True
        ws.column_dimensions[col_letter].width = col_width
        for r in ws.range("%s1:%s%d" % (col_letter, col_letter, len(table)+1)):
            for c in r:
                c.style.alignment.horizontal = Alignment.HORIZONTAL_LEFT
    ws._set_auto_filter("A1:%s1" % (col_letter))  #uses last col_letter value from loop above
    for row_idx in range(0, len(table)):
        for col_idx in range(0, len(table[row_idx])):
            ws.cell("%s%s" % (get_column_letter(col_idx+1), row_idx+2)).value = table[row_idx][col_idx]
    wb.save('%s' % (filename))

